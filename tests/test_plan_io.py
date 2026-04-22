from __future__ import annotations

import os
import unittest
from pathlib import Path
from unittest.mock import Mock

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from openpyxl import load_workbook
from PyQt5.QtWidgets import QApplication

from modbus_gui.main_window import ModbusMainWindow
from modbus_gui.models import RegisterFormat, RegisterValueType


FIXTURE_PATH = Path(__file__).parent / "artifacts" / "sample_plan_runtime.xlsx"
OUTPUT_PATH = Path(__file__).parent / "artifacts" / "roundtrip.xlsx"


class PlanIoTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.app = QApplication.instance() or QApplication([])
        cls._create_fixture()

    @classmethod
    def _create_fixture(cls) -> None:
        window = ModbusMainWindow()
        try:
            window.host_input.setText("127.0.0.1")
            window.port_input.setValue(502)
            window.slave_id_input.setValue(1)
            window.keepalive_input.setValue(20)
            window.pt1_p_input.setValue(150)
            window.pt1_p_register_input.setValue(1200)
            window.pt1_q_input.setValue(250)
            window.pt1_q_register_input.setValue(1300)
            window._set_combo_value(window.register_format_combo, RegisterFormat.LITTLE_BYTE_SWAP.value)

            window.channel_label_inputs[0].setText("Freigabe")
            window.channel_label_inputs[1].setText("Drehzahl")
            window.channel_label_inputs[2].setText("Wirkleistung")
            window.channel_label_inputs[3].setText("Blindleistung")
            window.register_inputs[0].setValue(0)
            window.register_inputs[1].setValue(1)
            window.register_inputs[2].setValue(3)
            window.register_inputs[3].setValue(5)
            window._set_combo_value(window.type_inputs[0], RegisterValueType.INT16.value)
            window._set_combo_value(window.type_inputs[1], RegisterValueType.UINT32.value)
            window._set_combo_value(window.type_inputs[2], RegisterValueType.FLOAT32.value)
            window._set_combo_value(window.type_inputs[3], RegisterValueType.FLOAT64.value)
            window.start_value_inputs[0].setText("1")
            window.start_value_inputs[2].setText("12.5")

            window.row_checkboxes[0].setChecked(True)
            window.row_checkboxes[2].setChecked(True)
            window._set_item_text(0, 2, "12")
            window._set_item_text(0, 3, "345")
            window._set_item_text(0, 4, "1.25")
            window._set_item_text(0, 5, "9.875")
            window._set_item_text(0, 6, "5")
            window._set_item_text(2, 2, "-7")
            window._set_item_text(2, 3, "65000")
            window._set_item_text(2, 4, "20.5")
            window._set_item_text(2, 5, "100.125")
            window._set_item_text(2, 6, "12")

            FIXTURE_PATH.parent.mkdir(parents=True, exist_ok=True)
            window._save_to_excel(FIXTURE_PATH)
            window._set_dirty(False)
        finally:
            window.close()

    def setUp(self) -> None:
        self.window = ModbusMainWindow()

    def tearDown(self) -> None:
        self.window._set_dirty(False)
        self.window.close()

    def test_load_fixture_populates_gui(self) -> None:
        self.window._load_from_excel(FIXTURE_PATH)

        self.assertEqual(self.window.host_input.text(), "127.0.0.1")
        self.assertEqual(self.window.port_input.value(), 502)
        self.assertEqual(self.window.slave_id_input.value(), 1)
        self.assertEqual(self.window.keepalive_input.value(), 20)
        self.assertEqual(self.window.pt1_p_input.value(), 150)
        self.assertEqual(self.window.pt1_p_register_input.value(), 1200)
        self.assertEqual(self.window.pt1_q_input.value(), 250)
        self.assertEqual(self.window.pt1_q_register_input.value(), 1300)
        self.assertEqual(self.window.register_format_combo.currentData(), RegisterFormat.LITTLE_BYTE_SWAP)

        self.assertEqual(self.window.channel_label_inputs[0].text(), "Freigabe")
        self.assertEqual(self.window.channel_label_inputs[1].text(), "Drehzahl")
        self.assertEqual(self.window.channel_label_inputs[2].text(), "Wirkleistung")
        self.assertEqual(self.window.channel_label_inputs[3].text(), "Blindleistung")
        self.assertEqual(self.window.register_inputs[0].value(), 0)
        self.assertEqual(self.window.register_inputs[1].value(), 1)
        self.assertEqual(self.window.register_inputs[2].value(), 3)
        self.assertEqual(self.window.register_inputs[3].value(), 5)
        self.assertEqual(self.window.type_inputs[0].currentData(), RegisterValueType.INT16)
        self.assertEqual(self.window.type_inputs[1].currentData(), RegisterValueType.UINT32)
        self.assertEqual(self.window.type_inputs[2].currentData(), RegisterValueType.FLOAT32)
        self.assertEqual(self.window.type_inputs[3].currentData(), RegisterValueType.FLOAT64)
        self.assertEqual(self.window.start_value_inputs[0].text(), "1")
        self.assertEqual(self.window.start_value_inputs[2].text(), "12.5")

        self.assertTrue(self.window.row_checkboxes[0].isChecked())
        self.assertFalse(self.window.row_checkboxes[1].isChecked())
        self.assertTrue(self.window.row_checkboxes[2].isChecked())
        self.assertEqual(self.window._item_text(0, 2), "12")
        self.assertEqual(self.window._item_text(0, 3), "345")
        self.assertEqual(self.window._item_text(0, 4), "1.25")
        self.assertEqual(self.window._item_text(0, 5), "9.875")
        self.assertEqual(self.window._item_text(0, 6), "5")
        self.assertEqual(self.window._item_text(2, 2), "-7")
        self.assertEqual(self.window._item_text(2, 3), "65000")
        self.assertEqual(self.window._item_text(2, 4), "20.5")
        self.assertEqual(self.window._item_text(2, 5), "100.125")
        self.assertEqual(self.window._item_text(2, 6), "12")

    def test_save_and_reload_roundtrip_preserves_values(self) -> None:
        self.window._load_from_excel(FIXTURE_PATH)
        self.window.host_input.setText("10.0.0.77")
        self.window.keepalive_input.setValue(30)
        self.window.pt1_p_input.setValue(500)
        self.window.pt1_p_register_input.setValue(2222)
        self.window.pt1_q_register_input.setValue(3333)
        self.window.channel_label_inputs[2].setText("Active Power")
        self.window.start_value_inputs[1].setText("99")
        self.window.row_checkboxes[1].setChecked(True)
        self.window._set_item_text(1, 2, "88")
        self.window._set_item_text(1, 3, "99")
        self.window._set_item_text(1, 4, "7.5")
        self.window._set_item_text(1, 5, "8.5")
        self.window._set_item_text(1, 6, "15")

        OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
        self.window._save_to_excel(OUTPUT_PATH)
        workbook = load_workbook(OUTPUT_PATH)

        connection_sheet = workbook["Verbindung"]
        connection = {
            str(row[0].value): row[1].value
            for row in connection_sheet.iter_rows(min_row=2, max_col=2)
            if row[0].value is not None
        }
        channel_sheet = workbook["Kanaele"]
        saved_channels = list(channel_sheet.iter_rows(min_row=2, max_col=5, values_only=True))
        plan_sheet = workbook["Testplan"]
        saved_rows = list(plan_sheet.iter_rows(min_row=2, max_col=7, values_only=True))

        self.assertEqual(connection["host"], "10.0.0.77")
        self.assertEqual(connection["keepalive_interval_seconds"], 30)
        self.assertEqual(connection["pt1_p_ms"], 500)
        self.assertEqual(connection["pt1_p_start_register"], 2222)
        self.assertEqual(connection["pt1_q_start_register"], 3333)
        self.assertEqual(saved_channels[2][1], "Active Power")
        self.assertEqual(saved_channels[1][4], "99")
        self.assertTrue(saved_rows[1][0])
        self.assertEqual(list(saved_rows[1][2:6]), ["88", "99", "7.5", "8.5"])
        self.assertEqual(str(saved_rows[1][6]), "15")

    def test_copy_first_stage_time_applies_to_active_rows(self) -> None:
        self.window.row_checkboxes[0].setChecked(True)
        self.window.row_checkboxes[1].setChecked(True)
        self.window.row_checkboxes[2].setChecked(True)
        self.window._set_item_text(0, self.window.COLUMN_DURATION, "7")
        self.window._set_item_text(1, self.window.COLUMN_DURATION, "2")
        self.window._set_item_text(2, self.window.COLUMN_DURATION, "9")

        self.window._copy_first_stage_time_to_active_rows()

        self.assertEqual(self.window._item_text(0, self.window.COLUMN_DURATION), "7")
        self.assertEqual(self.window._item_text(1, self.window.COLUMN_DURATION), "7")
        self.assertEqual(self.window._item_text(2, self.window.COLUMN_DURATION), "7")

    def test_workspace_tabs_separate_key_areas(self) -> None:
        self.assertEqual(self.window.workspace_tabs.count(), 4)
        self.assertEqual(
            [self.window.workspace_tabs.tabText(index) for index in range(self.window.workspace_tabs.count())],
            ["Testplan", "Kanaele", "Automatisierung", "Protokoll"],
        )

    def test_current_plan_can_be_added_to_automation_queue(self) -> None:
        self.window.current_file_path = Path("C:/Tests/Basislauf.xlsx")
        expected_source = str(self.window.current_file_path)

        self.window._on_add_current_plan_to_automation_clicked()

        self.assertTrue(self.window.automation_row_checkboxes[0].isChecked())
        self.assertEqual(self.window._automation_item_text(0, self.window.AUTOMATION_COLUMN_NAME), "Basislauf")
        self.assertEqual(
            self.window._automation_item_text(0, self.window.AUTOMATION_COLUMN_SOURCE),
            expected_source,
        )
        self.assertEqual(self.window._automation_item_text(0, self.window.AUTOMATION_COLUMN_REPEAT), "1")

    def test_save_and_load_roundtrip_preserves_automation_queue(self) -> None:
        self.window._append_automation_entry(
            name="Basislauf",
            source="C:/Tests/Basislauf.xlsx",
            repeat_count="3",
            note="Abgleich mit Referenzgeraet",
        )
        self.window._append_automation_entry(
            name="Variantenlauf",
            source="C:/Tests/Variante_A.xlsx",
            repeat_count="2",
            note="Warmstart",
        )

        OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
        self.window._save_to_excel(OUTPUT_PATH)
        workbook = load_workbook(OUTPUT_PATH)

        self.assertIn("Automatisierung", workbook.sheetnames)
        automation_sheet = workbook["Automatisierung"]
        saved_jobs = list(automation_sheet.iter_rows(min_row=2, max_col=5, values_only=True))
        self.assertEqual(saved_jobs[0], (True, "Basislauf", "C:/Tests/Basislauf.xlsx", "3", "Abgleich mit Referenzgeraet"))
        self.assertEqual(saved_jobs[1], (True, "Variantenlauf", "C:/Tests/Variante_A.xlsx", "2", "Warmstart"))

        reloaded_window = ModbusMainWindow()
        try:
            reloaded_window._load_from_excel(OUTPUT_PATH)
            self.assertTrue(reloaded_window.automation_row_checkboxes[0].isChecked())
            self.assertEqual(
                reloaded_window._automation_item_text(0, reloaded_window.AUTOMATION_COLUMN_NAME),
                "Basislauf",
            )
            self.assertEqual(
                reloaded_window._automation_item_text(0, reloaded_window.AUTOMATION_COLUMN_SOURCE),
                "C:/Tests/Basislauf.xlsx",
            )
            self.assertEqual(
                reloaded_window._automation_item_text(1, reloaded_window.AUTOMATION_COLUMN_NOTE),
                "Warmstart",
            )
            self.assertEqual(
                reloaded_window._automation_item_text(0, reloaded_window.AUTOMATION_COLUMN_STATUS),
                "Bereit",
            )
        finally:
            reloaded_window._set_dirty(False)
            reloaded_window.close()

    def test_automation_run_advances_queue_and_restores_editor_state(self) -> None:
        self.window.channel_label_inputs[0].setText("Originalzustand")
        self.window.current_file_path = Path("C:/Tests/Serienmaster.xlsx")
        self.window._set_dirty(True)
        self.window._append_automation_entry(
            name="Plan A",
            source=str(FIXTURE_PATH),
            repeat_count="1",
            note="Erster Durchgang",
        )
        self.window._append_automation_entry(
            name="Plan B",
            source=str(FIXTURE_PATH),
            repeat_count="2",
            note="Zweiter Durchgang",
        )
        self.window.modbus_service._client = Mock()
        self.window.modbus_service._settings = self.window._current_settings()
        self.window.sequence_controller.start = Mock()

        self.window._on_automation_start_clicked()

        self.assertTrue(self.window.series_controller.is_running)
        self.assertEqual(self.window.sequence_controller.start.call_count, 1)
        self.assertEqual(self.window.current_file_path, Path("C:/Tests/Serienmaster.xlsx"))
        self.assertEqual(self.window.workspace_tabs.tabText(self.window.workspace_tabs.currentIndex()), "Testplan")
        self.assertEqual(
            self.window._automation_item_text(0, self.window.AUTOMATION_COLUMN_STATUS),
            "Laeuft 1/1",
        )
        self.assertEqual(
            self.window._automation_item_text(1, self.window.AUTOMATION_COLUMN_STATUS),
            "Wartet",
        )
        self.assertEqual(self.window._automation_reports[0].started_runs, 1)
        self.assertIn("gestartet", self.window._automation_reports[0].last_message.lower())

        self.window._on_sequence_finished()
        self.assertEqual(self.window.sequence_controller.start.call_count, 2)
        self.assertEqual(
            self.window._automation_item_text(0, self.window.AUTOMATION_COLUMN_STATUS),
            "Fertig",
        )
        self.assertEqual(
            self.window._automation_item_text(1, self.window.AUTOMATION_COLUMN_STATUS),
            "Laeuft 1/2",
        )
        self.assertEqual(self.window._automation_reports[0].successful_runs, 1)

        self.window._on_sequence_finished()
        self.assertEqual(self.window.sequence_controller.start.call_count, 3)
        self.assertEqual(
            self.window._automation_item_text(1, self.window.AUTOMATION_COLUMN_STATUS),
            "Laeuft 2/2",
        )

        self.window._on_sequence_finished()
        self.assertFalse(self.window.series_controller.is_running)
        self.assertEqual(self.window.sequence_controller.start.call_count, 3)
        self.assertEqual(self.window.channel_label_inputs[0].text(), "Originalzustand")
        self.assertEqual(self.window.current_file_path, Path("C:/Tests/Serienmaster.xlsx"))
        self.assertTrue(self.window._has_unsaved_changes)
        self.assertEqual(
            self.window._automation_item_text(0, self.window.AUTOMATION_COLUMN_STATUS),
            "Fertig",
        )
        self.assertEqual(
            self.window._automation_item_text(1, self.window.AUTOMATION_COLUMN_STATUS),
            "Fertig",
        )
        self.assertEqual(self.window._automation_reports[1].successful_runs, 2)
        self.assertGreaterEqual(len(self.window._automation_reports[1].history), 3)
        self.assertEqual(self.window.automation_result_status_value.text(), "Fertig")
        self.assertEqual(self.window.automation_result_success_value.text(), "2")
        self.assertIn("erfolgreich", self.window.automation_result_message_value.text().lower())

    def test_connect_can_retry_after_failure(self) -> None:
        first_error = RuntimeError("slave antwortet nicht")
        self.window.modbus_service.connect = Mock(side_effect=[first_error, None])
        self.window.modbus_service.disconnect = Mock()
        self.window._show_retry_dialog = Mock(side_effect=[True])

        self.window._on_connect_clicked()

        self.assertEqual(self.window.modbus_service.connect.call_count, 2)
        self.window.modbus_service.disconnect.assert_called_once()
        self.window._show_retry_dialog.assert_called_once()


if __name__ == "__main__":
    unittest.main()
