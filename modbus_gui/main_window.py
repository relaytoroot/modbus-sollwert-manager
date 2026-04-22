from __future__ import annotations

import sys
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QCloseEvent, QDoubleValidator, QColor, QIcon, QKeySequence, QPixmap
from PyQt5.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QCheckBox,
    QComboBox,
    QFileDialog,
    QFormLayout,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QPlainTextEdit,
    QPushButton,
    QSpinBox,
    QStyledItemDelegate,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from .modbus_service import ModbusService
from .models import (
    AutomationJob,
    ChannelConfig,
    ChannelWrite,
    ConnectionSettings,
    RegisterFormat,
    RegisterValueType,
    StageExecution,
    STATUS_CONNECTED,
    STATUS_DISCONNECTED,
    STATUS_ERROR,
    STATUS_FINISHED,
    STATUS_PAUSED,
    STATUS_RUNNING,
)
from .app_info import (
    APP_AUTHOR,
    APP_COMPANY,
    APP_NAME,
    APP_VERSION,
    HEADER_LOGO_FILE,
    ICON_FILE,
    resource_path,
)
from .sequence_controller import SequenceController
from .series_controller import SeriesController
from .value_encoder import ValueEncoder, ValueEncodingError


class NumericItemDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, index):  # type: ignore[override]
        editor = QLineEdit(parent)
        validator = QDoubleValidator(editor)
        validator.setNotation(QDoubleValidator.StandardNotation)
        validator.setDecimals(8)
        editor.setValidator(validator)
        editor.setAlignment(Qt.AlignCenter)
        return editor


class PlanTableWidget(QTableWidget):
    def keyPressEvent(self, event):  # type: ignore[override]
        if event.matches(QKeySequence.Copy):
            self._copy_selection()
            return
        if event.matches(QKeySequence.Paste):
            self._paste_selection()
            return
        if event.key() in {Qt.Key_Delete, Qt.Key_Backspace}:
            self._clear_selection()
            return
        super().keyPressEvent(event)

    def _copy_selection(self) -> None:
        ranges = self.selectedRanges()
        if not ranges:
            return
        selected = ranges[0]
        rows: list[str] = []
        for row in range(selected.topRow(), selected.bottomRow() + 1):
            values: list[str] = []
            for column in range(selected.leftColumn(), selected.rightColumn() + 1):
                item = self.item(row, column)
                values.append("" if item is None else item.text())
            rows.append("\t".join(values))
        QApplication.clipboard().setText("\n".join(rows))

    def _paste_selection(self) -> None:
        text = QApplication.clipboard().text()
        if not text:
            return
        start_row = max(0, self.currentRow())
        start_column = max(2, self.currentColumn())
        for row_offset, line in enumerate(text.splitlines()):
            for column_offset, value in enumerate(line.split("\t")):
                row = start_row + row_offset
                column = start_column + column_offset
                if row >= self.rowCount() or column >= self.columnCount() or column < 2:
                    continue
                item = self.item(row, column)
                if item is not None and (item.flags() & Qt.ItemIsEditable):
                    item.setText(value)

    def _clear_selection(self) -> None:
        for item in self.selectedItems():
            if item.column() >= 2 and (item.flags() & Qt.ItemIsEditable):
                item.setText("")


class ModbusMainWindow(QMainWindow):
    THEMES = {
        "light": {
            "window_bg": "#f4f7fa",
            "text_primary": "#1d2733",
            "text_secondary": "#607080",
            "group_bg": "#ffffff",
            "border": "#d4dde6",
            "button_bg": "#f6f8fb",
            "button_hover": "#eef3f7",
            "button_disabled_bg": "#eef2f5",
            "button_disabled_text": "#96a1ac",
            "input_bg": "#ffffff",
            "input_border": "#cad5df",
            "table_bg": "#ffffff",
            "table_alt_bg": "#f7f9fb",
            "table_grid": "#e4e9ef",
            "table_header_bg": "#eef3f7",
            "table_selection_bg": "#dfe9f4",
            "table_selection_text": "#13202c",
            "default_row_bg": "#ffffff",
            "inactive_row_bg": "#f3f5f7",
            "active_row_bg": "#e8f0f7",
            "active_row_text": "#13202c",
            "inactive_row_text": "#667585",
            "author_text": "#6b7885",
            "status_styles": {
                STATUS_DISCONNECTED: "background:#e3e8ed;color:#2f4354;",
                STATUS_CONNECTED: "background:#d9ebe1;color:#24543a;",
                STATUS_RUNNING: "background:#f3e5c8;color:#73511d;",
                STATUS_PAUSED: "background:#e7edf4;color:#37526b;",
                STATUS_ERROR: "background:#f3d9d7;color:#7d2e28;",
                STATUS_FINISHED: "background:#d8e7f4;color:#234d70;",
            },
        },
        "dark": {
            "window_bg": "#1f252c",
            "text_primary": "#e6edf3",
            "text_secondary": "#a9b5c1",
            "group_bg": "#28303a",
            "border": "#3b4652",
            "button_bg": "#313b46",
            "button_hover": "#394652",
            "button_disabled_bg": "#2b333d",
            "button_disabled_text": "#7e8a97",
            "input_bg": "#202830",
            "input_border": "#44505d",
            "table_bg": "#242c34",
            "table_alt_bg": "#20272f",
            "table_grid": "#37414c",
            "table_header_bg": "#313b46",
            "table_selection_bg": "#35516b",
            "table_selection_text": "#f2f7fb",
            "default_row_bg": "#28303a",
            "inactive_row_bg": "#232a33",
            "active_row_bg": "#344658",
            "active_row_text": "#f1f6fb",
            "inactive_row_text": "#95a3b1",
            "author_text": "#94a3b0",
            "status_styles": {
                STATUS_DISCONNECTED: "background:#3a4754;color:#e7edf2;",
                STATUS_CONNECTED: "background:#2f5644;color:#e8f4ed;",
                STATUS_RUNNING: "background:#6a5324;color:#fff4db;",
                STATUS_PAUSED: "background:#3d5266;color:#eef5fb;",
                STATUS_ERROR: "background:#6a3734;color:#ffe9e7;",
                STATUS_FINISHED: "background:#31506a;color:#e7f3fb;",
            },
        },
    }
    DEFAULT_WINDOW_TITLE = APP_NAME
    COLUMN_ACTIVE = 0
    COLUMN_STAGE = 1
    COLUMN_VALUE_START = 2
    COLUMN_DURATION = 6
    CHANNEL_COUNT = 4
    AUTOMATION_COLUMN_ACTIVE = 0
    AUTOMATION_COLUMN_NAME = 1
    AUTOMATION_COLUMN_SOURCE = 2
    AUTOMATION_COLUMN_REPEAT = 3
    AUTOMATION_COLUMN_NOTE = 4
    AUTOMATION_ROW_COUNT = 20

    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle(self.DEFAULT_WINDOW_TITLE)
        self.resize(1600, 980)
        self.logo_path = resource_path(ICON_FILE)
        self.header_logo_path = resource_path(HEADER_LOGO_FILE)

        self.modbus_service = ModbusService()
        self.sequence_controller = SequenceController(self.modbus_service, self)
        self.series_controller = SeriesController(self._start_automation_job, self)
        self.keepalive_timer = QTimer(self)
        self.keepalive_timer.timeout.connect(self._on_keepalive_tick)
        self.current_file_path: Path | None = None
        self._has_unsaved_changes = False
        self._automation_snapshot_bytes: bytes | None = None
        self._automation_snapshot_file_path: Path | None = None
        self._automation_snapshot_dirty = False
        self.row_checkboxes: list[QCheckBox] = []
        self.automation_row_checkboxes: list[QCheckBox] = []
        self.register_inputs: list[QSpinBox] = []
        self.channel_label_inputs: list[QLineEdit] = []
        self.type_inputs: list[QComboBox] = []
        self.start_value_inputs: list[QLineEdit] = []
        self.send_value_buttons: list[QPushButton] = []
        self.current_highlight_row = -1
        self._table_ui_update_in_progress = False
        self._automation_ui_update_in_progress = False
        self.current_theme = "light"
        self.current_status_key = STATUS_DISCONNECTED
        self.current_status_text = "Getrennt"
        self.default_row_color = QColor("#ffffff")
        self.inactive_row_color = QColor("#f1f3f5")
        self.active_stage_color = QColor("#d9ecff")
        self.active_row_text_color = QColor("#13202c")
        self.inactive_row_text_color = QColor("#667585")

        self.host_input = QLineEdit("127.0.0.1")
        self.port_input = QSpinBox()
        self.port_input.setRange(1, 65535)
        self.port_input.setValue(502)
        self.slave_id_input = QSpinBox()
        self.slave_id_input.setRange(1, 247)
        self.slave_id_input.setValue(1)
        self.keepalive_input = QSpinBox()
        self.keepalive_input.setRange(1, 3600)
        self.keepalive_input.setSuffix(" s")
        self.keepalive_input.setValue(20)

        self.register_format_combo = QComboBox()
        for register_format in RegisterFormat:
            self.register_format_combo.addItem(register_format.label, register_format)
        self.register_format_combo.setMaximumWidth(180)
        self.theme_combo = QComboBox()
        self.theme_combo.addItem("Hell", "light")
        self.theme_combo.addItem("Dunkel", "dark")
        self.theme_combo.setMaximumWidth(120)
        self.pt1_p_input = QSpinBox()
        self.pt1_p_input.setRange(0, 1_000_000)
        self.pt1_p_input.setSuffix(" ms")
        self.pt1_p_input.setMaximumWidth(92)
        self.pt1_p_register_input = QSpinBox()
        self.pt1_p_register_input.setRange(0, 65535)
        self.pt1_p_register_input.setMaximumWidth(84)
        self.pt1_q_input = QSpinBox()
        self.pt1_q_input.setRange(0, 1_000_000)
        self.pt1_q_input.setSuffix(" ms")
        self.pt1_q_input.setMaximumWidth(92)
        self.pt1_q_register_input = QSpinBox()
        self.pt1_q_register_input.setRange(0, 65535)
        self.pt1_q_register_input.setMaximumWidth(84)

        for index in range(4):
            register_input = QSpinBox()
            register_input.setRange(0, 65535)
            register_input.setValue((0, 1, 3, 5)[index])
            channel_label_input = QLineEdit()
            channel_label_input.setPlaceholderText("z. B. Wirkleistung")
            channel_label_input.setMaximumWidth(156)
            type_input = QComboBox()
            for value_type in RegisterValueType:
                type_input.addItem(value_type.label, value_type)
            type_input.setCurrentIndex(type_input.findData(RegisterValueType.FLOAT32))
            start_value_input = QLineEdit()
            start_value_input.setPlaceholderText("Startwert")
            start_value_input.setMaximumWidth(88)
            send_value_button = QPushButton("Send")
            send_value_button.setMaximumWidth(64)
            register_input.setMaximumWidth(84)
            type_input.setMaximumWidth(126)
            self.register_inputs.append(register_input)
            self.channel_label_inputs.append(channel_label_input)
            self.type_inputs.append(type_input)
            self.start_value_inputs.append(start_value_input)
            self.send_value_buttons.append(send_value_button)

        self.connect_button = QPushButton("Verbinden")
        self.disconnect_button = QPushButton("Trennen")
        self.start_button = QPushButton("Start")
        self.pause_button = QPushButton("Pause")
        self.next_stage_button = QPushButton("Naechste Stufe")
        self.stop_button = QPushButton("Stopp")
        self.save_button = QPushButton("Speichern unter")
        self.load_button = QPushButton("Laden")
        self.copy_stage_time_button = QPushButton("Zeit aus Stufe 1")
        self.copy_stage_time_button.setToolTip(
            "Uebernimmt die Zeit aus Stufe 1 in alle anderen aktiven Stufen."
        )
        self.copy_stage_time_button.setMaximumWidth(150)
        self.disconnect_button.setEnabled(False)
        self.pause_button.setEnabled(False)
        self.next_stage_button.setEnabled(False)
        self.stop_button.setEnabled(False)
        self.copy_stage_time_button.setEnabled(False)

        self.logo_label = QLabel()
        self._configure_logo()
        self.author_label = QLabel(
            f"{APP_COMPANY} | {APP_NAME} | Version {APP_VERSION} | {APP_AUTHOR}"
        )
        self.author_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)

        self.status_badge = QLabel("Getrennt")
        self.status_badge.setAlignment(Qt.AlignCenter)
        self.status_badge.setMinimumWidth(120)
        self.current_stage_value = QLabel("-")
        self.stage_remaining_value = QLabel("00:00:00")
        self.total_remaining_value = QLabel("00:00:00")
        self.total_time_value = QLabel("00:00:00")
        self.last_write_value = QLabel("Noch keine Schreibaktion")
        self.address_info = QLabel("Registeradressen sind nullbasiert.")

        self.log_panel = QPlainTextEdit()
        self.log_panel.setReadOnly(True)
        self.log_panel.setMaximumBlockCount(1000)
        self.log_panel.setMinimumHeight(120)
        self.workspace_tabs = QTabWidget()
        self.workspace_tabs.setObjectName("workspaceTabs")
        self.automation_summary_label = QLabel("Noch keine Serienjobs geplant.")
        self.automation_summary_label.setWordWrap(True)
        self.automation_start_button = QPushButton("Serienlauf starten")
        self.automation_add_current_button = QPushButton("Aktuellen Plan uebernehmen")
        self.automation_add_file_button = QPushButton("Plan aus Datei")
        self.automation_remove_button = QPushButton("Auswahl entfernen")
        self.automation_move_up_button = QPushButton("Nach oben")
        self.automation_move_down_button = QPushButton("Nach unten")
        self.automation_clear_button = QPushButton("Liste leeren")
        self.automation_table = QTableWidget(self.AUTOMATION_ROW_COUNT, 5)

        self.test_table = PlanTableWidget(40, 7)
        editable_columns = [
            *range(self.COLUMN_VALUE_START, self.COLUMN_VALUE_START + self.CHANNEL_COUNT),
            self.COLUMN_DURATION,
        ]
        for column in editable_columns:
            self.test_table.setItemDelegateForColumn(column, NumericItemDelegate(self.test_table))

        self._configure_table()
        self._configure_automation_table()
        self._build_ui()
        self._apply_styles()
        self._connect_signals()
        self._set_status(STATUS_DISCONNECTED, "Getrennt")
        self._apply_row_visuals()
        self._refresh_summary()
        self._refresh_automation_summary()
        self._update_automation_buttons()
        self._update_manual_write_buttons()
        self._append_log("Anwendung gestartet. Bitte Verbindung pruefen und Testplan laden oder erstellen.")

    def _build_ui(self) -> None:
        central = QWidget(self)
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(18, 18, 18, 18)
        root.setSpacing(14)

        top_cards = QHBoxLayout()
        top_cards.addWidget(self._build_connection_group(), 3)
        top_cards.addWidget(self._build_format_group(), 2)
        top_cards.addWidget(self._build_control_group(), 3)
        root.addLayout(top_cards)
        root.addWidget(self._build_workspace_tabs(), 1)
        root.addWidget(self._build_footer())

    def _build_workspace_tabs(self) -> QTabWidget:
        self.workspace_tabs.setDocumentMode(True)
        self.workspace_tabs.addTab(self._build_testplan_tab(), "Testplan")
        self.workspace_tabs.addTab(self._build_channel_tab(), "Kanaele")
        self.workspace_tabs.addTab(self._build_automation_tab(), "Automatisierung")
        self.workspace_tabs.addTab(self._build_log_tab(), "Protokoll")
        return self.workspace_tabs

    def _build_testplan_tab(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(0, 10, 0, 0)
        layout.setSpacing(12)
        layout.addWidget(self._build_table_group(), 1)
        return tab

    def _build_channel_tab(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(0, 10, 0, 0)
        layout.setSpacing(12)
        layout.addWidget(self._build_channel_group(), 0, Qt.AlignTop)
        layout.addStretch(1)
        return tab

    def _build_automation_tab(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(0, 10, 0, 0)
        layout.setSpacing(12)
        layout.addWidget(self._build_automation_overview_group())
        layout.addWidget(self._build_automation_queue_group(), 1)
        layout.addWidget(self._build_automation_hint_group())
        return tab

    def _build_automation_overview_group(self) -> QGroupBox:
        group = QGroupBox("Serienplanung")
        layout = QVBoxLayout(group)
        layout.setSpacing(10)

        button_row = QHBoxLayout()
        button_row.addWidget(self.automation_start_button)
        button_row.addWidget(self.automation_add_current_button)
        button_row.addWidget(self.automation_add_file_button)
        button_row.addWidget(self.automation_remove_button)
        button_row.addWidget(self.automation_move_up_button)
        button_row.addWidget(self.automation_move_down_button)
        button_row.addWidget(self.automation_clear_button)
        button_row.addStretch(1)
        layout.addLayout(button_row)
        layout.addWidget(self.automation_summary_label)
        return group

    def _build_automation_queue_group(self) -> QGroupBox:
        group = QGroupBox("Serienliste")
        layout = QVBoxLayout(group)
        layout.addWidget(self.automation_table)
        return group

    def _build_automation_hint_group(self) -> QGroupBox:
        group = QGroupBox("Hinweis")
        layout = QVBoxLayout(group)
        info_text = QLabel(
            "Hier bereitest du bereits mehrere Testplaene, Wiederholungen und Kommentare "
            "fuer einen spaeteren Serienlauf vor. Die eigentliche automatische Ausfuehrung "
            "bauen wir im naechsten Schritt auf dieser Struktur auf."
        )
        info_text.setWordWrap(True)
        info_text.setAlignment(Qt.AlignTop | Qt.AlignLeft)
        layout.addWidget(info_text)
        return group

    def _build_log_tab(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(0, 10, 0, 0)
        layout.setSpacing(12)
        layout.addWidget(self._build_log_group(), 1)
        return tab

    def _build_connection_group(self) -> QGroupBox:
        group = QGroupBox("Verbindung")
        layout = QGridLayout(group)

        logo_col = QVBoxLayout()
        logo_col.addWidget(self.logo_label, 0, Qt.AlignLeft | Qt.AlignTop)
        logo_col.addStretch(1)
        layout.addLayout(logo_col, 0, 0, 2, 1)

        form = QFormLayout()
        form.addRow("IP-Adresse", self.host_input)
        form.addRow("Port", self.port_input)
        form.addRow("Slave-ID", self.slave_id_input)
        form.addRow("Keepalive", self.keepalive_input)
        layout.addLayout(form, 0, 1, 2, 1)

        button_col = QVBoxLayout()
        button_col.addWidget(self.connect_button)
        button_col.addWidget(self.disconnect_button)
        button_col.addStretch(1)
        layout.addLayout(button_col, 0, 2)

        status_col = QVBoxLayout()
        status_col.addWidget(QLabel("Status"))
        status_col.addWidget(self.status_badge)
        status_col.addWidget(self.address_info)
        status_col.addStretch(1)
        layout.addLayout(status_col, 0, 3)
        layout.setColumnStretch(1, 1)
        return group

    def _build_format_group(self) -> QGroupBox:
        group = QGroupBox("Datenformat")
        form = QFormLayout(group)
        form.setLabelAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        form.setFormAlignment(Qt.AlignTop)
        form.setHorizontalSpacing(10)
        group.setMaximumWidth(250)
        form.addRow("Registerformat", self.register_format_combo)
        form.addRow("Darstellung", self.theme_combo)
        return group

    def _build_control_group(self) -> QGroupBox:
        group = QGroupBox("Ablauf")
        layout = QGridLayout(group)
        layout.addWidget(self.start_button, 0, 0)
        layout.addWidget(self.pause_button, 0, 1)
        layout.addWidget(self.next_stage_button, 0, 2)
        layout.addWidget(self.stop_button, 0, 3)
        layout.addWidget(self.save_button, 1, 0)
        layout.addWidget(self.load_button, 1, 1)
        layout.addWidget(QLabel("Aktuelle Stufe"), 0, 4)
        layout.addWidget(self.current_stage_value, 0, 5)
        layout.addWidget(QLabel("Restzeit Stufe"), 1, 4)
        layout.addWidget(self.stage_remaining_value, 1, 5)
        layout.addWidget(QLabel("Restzeit gesamt"), 0, 6)
        layout.addWidget(self.total_remaining_value, 0, 7)
        layout.addWidget(QLabel("Gesamtzeit"), 1, 6)
        layout.addWidget(self.total_time_value, 1, 7)
        layout.addWidget(QLabel("Letzte Rueckmeldung"), 2, 0, 1, 4)
        layout.addWidget(self.last_write_value, 2, 4, 1, 4)
        return group

    def _build_channel_group(self) -> QGroupBox:
        group = QGroupBox("Kanalzuordnung")
        group.setMaximumHeight(190)
        outer_layout = QHBoxLayout(group)
        outer_layout.setContentsMargins(12, 12, 12, 12)
        outer_layout.setSpacing(12)

        mapping_widget = QWidget()
        mapping_widget.setMaximumWidth(600)
        layout = QGridLayout(mapping_widget)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setHorizontalSpacing(6)
        layout.setVerticalSpacing(4)
        layout.addWidget(QLabel("Kanal"), 0, 0)
        layout.addWidget(QLabel("Bezeichnung"), 0, 1)
        layout.addWidget(QLabel("Startregister"), 0, 2)
        layout.addWidget(QLabel("Datentyp"), 0, 3)
        layout.addWidget(QLabel("Startwert"), 0, 4)
        layout.addWidget(QLabel("Aktion"), 0, 5)
        for index, (channel_label_input, register_input, type_input, start_value_input, send_value_button) in enumerate(
            zip(self.channel_label_inputs, self.register_inputs, self.type_inputs, self.start_value_inputs, self.send_value_buttons),
            start=1,
        ):
            layout.addWidget(QLabel(f"S{index}"), index, 0)
            layout.addWidget(channel_label_input, index, 1)
            layout.addWidget(register_input, index, 2)
            layout.addWidget(type_input, index, 3)
            layout.addWidget(start_value_input, index, 4)
            layout.addWidget(send_value_button, index, 5)

        pt1_widget = QWidget()
        pt1_widget.setMaximumWidth(165)
        pt1_layout = QFormLayout(pt1_widget)
        pt1_layout.setContentsMargins(0, 0, 0, 0)
        pt1_layout.setLabelAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        pt1_layout.setHorizontalSpacing(6)
        pt1_layout.setVerticalSpacing(6)
        pt1_layout.addRow("PT1 P [ms]", self.pt1_p_input)
        pt1_layout.addRow("PT1 P Reg.", self.pt1_p_register_input)
        pt1_layout.addRow("PT1 Q [ms]", self.pt1_q_input)
        pt1_layout.addRow("PT1 Q Reg.", self.pt1_q_register_input)

        outer_layout.addWidget(mapping_widget, 0, Qt.AlignLeft | Qt.AlignTop)
        outer_layout.addWidget(pt1_widget, 0, Qt.AlignTop)
        outer_layout.addStretch(1)
        return group

    def _build_table_group(self) -> QGroupBox:
        group = QGroupBox("Testplan")
        layout = QVBoxLayout(group)
        helper_row = QHBoxLayout()
        helper_row.setContentsMargins(0, 0, 0, 0)
        helper_row.addStretch(1)
        helper_row.addWidget(self.copy_stage_time_button, 0, Qt.AlignRight)
        layout.addLayout(helper_row)
        layout.addWidget(self.test_table)
        return group

    def _build_log_group(self) -> QGroupBox:
        group = QGroupBox("Ereignisprotokoll")
        layout = QVBoxLayout(group)
        layout.addWidget(self.log_panel)
        return group

    def _build_footer(self) -> QWidget:
        footer = QWidget()
        layout = QHBoxLayout(footer)
        layout.setContentsMargins(2, 0, 2, 0)
        layout.addStretch(1)
        layout.addWidget(self.author_label, 0, Qt.AlignRight | Qt.AlignVCenter)
        return footer

    def _configure_logo(self) -> None:
        if self.logo_path.exists():
            pixmap = QPixmap(str(self.logo_path)).scaled(
                168,
                64,
                Qt.KeepAspectRatio,
                Qt.SmoothTransformation,
            )
            self.logo_label.setPixmap(pixmap)
            self.logo_label.setMinimumSize(168, 64)
            self.logo_label.setMaximumSize(168, 64)
            self.logo_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            window_icon_path = self.header_logo_path if self.header_logo_path.exists() else self.logo_path
            self.setWindowIcon(QIcon(str(window_icon_path)))
        else:
            self.logo_label.setText("")

    def _apply_styles(self) -> None:
        self.status_badge.setObjectName("statusBadge")
        self.author_label.setObjectName("authorLabel")
        self.copy_stage_time_button.setObjectName("secondaryButton")
        self.next_stage_button.setObjectName("secondaryButton")
        self.automation_remove_button.setObjectName("secondaryButton")
        self.automation_move_up_button.setObjectName("secondaryButton")
        self.automation_move_down_button.setObjectName("secondaryButton")
        self.automation_clear_button.setObjectName("secondaryButton")
        self._apply_theme(self.current_theme)

    def _apply_theme(self, theme_name: str) -> None:
        theme = self.THEMES.get(theme_name, self.THEMES["light"])
        self.current_theme = theme_name if theme_name in self.THEMES else "light"
        self.default_row_color = QColor(theme["default_row_bg"])
        self.inactive_row_color = QColor(theme["inactive_row_bg"])
        self.active_stage_color = QColor(theme["active_row_bg"])
        self.active_row_text_color = QColor(theme["active_row_text"])
        self.inactive_row_text_color = QColor(theme["inactive_row_text"])

        self.setStyleSheet(
            f"""
            QWidget {{
                font-size: 13px;
                color: {theme["text_primary"]};
                background: {theme["window_bg"]};
            }}
            QMainWindow {{
                background: {theme["window_bg"]};
            }}
            QGroupBox {{
                border: 1px solid {theme["border"]};
                border-radius: 10px;
                margin-top: 10px;
                padding-top: 12px;
                background: {theme["group_bg"]};
                font-weight: 600;
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                left: 12px;
                padding: 0 6px;
            }}
            QTabWidget::pane {{
                border: 1px solid {theme["border"]};
                border-radius: 10px;
                background: {theme["window_bg"]};
                margin-top: 10px;
            }}
            QTabBar::tab {{
                background: {theme["button_bg"]};
                border: 1px solid {theme["input_border"]};
                border-bottom: none;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                padding: 9px 16px;
                margin-right: 4px;
                min-width: 118px;
            }}
            QTabBar::tab:hover {{
                background: {theme["button_hover"]};
            }}
            QTabBar::tab:selected {{
                background: {theme["group_bg"]};
                color: {theme["text_primary"]};
            }}
            QPushButton {{
                min-height: 34px;
                border-radius: 8px;
                border: 1px solid {theme["input_border"]};
                padding: 0 14px;
                background: {theme["button_bg"]};
            }}
            QPushButton:hover {{
                background: {theme["button_hover"]};
            }}
            QPushButton:disabled {{
                color: {theme["button_disabled_text"]};
                background: {theme["button_disabled_bg"]};
            }}
            QLineEdit, QComboBox, QSpinBox, QPlainTextEdit {{
                border: 1px solid {theme["input_border"]};
                border-radius: 8px;
                padding: 6px 8px;
                background: {theme["input_bg"]};
                color: {theme["text_primary"]};
                selection-background-color: {theme["table_selection_bg"]};
                selection-color: {theme["table_selection_text"]};
            }}
            QTableWidget {{
                border: 1px solid {theme["border"]};
                background: {theme["table_bg"]};
                color: {theme["text_primary"]};
                gridline-color: {theme["table_grid"]};
                alternate-background-color: {theme["table_alt_bg"]};
                selection-background-color: {theme["table_selection_bg"]};
                selection-color: {theme["table_selection_text"]};
            }}
            QTableWidget::item:selected {{
                background: {theme["table_selection_bg"]};
                color: {theme["table_selection_text"]};
            }}
            QTableCornerButton::section {{
                background: {theme["table_header_bg"]};
                border: none;
                border-bottom: 1px solid {theme["border"]};
                border-right: 1px solid {theme["border"]};
            }}
            QHeaderView::section {{
                background: {theme["table_header_bg"]};
                color: {theme["text_primary"]};
                border: none;
                border-bottom: 1px solid {theme["border"]};
                padding: 8px;
                font-weight: 600;
            }}
            QLabel#statusBadge {{
                border-radius: 12px;
                padding: 8px 12px;
                font-weight: 700;
            }}
            QPushButton#secondaryButton {{
                padding: 0 10px;
            }}
            QLabel#authorLabel {{
                color: {theme["author_text"]};
                font-size: 9px;
                font-weight: 400;
                padding: 1px 4px;
            }}
            """
        )
        self._set_status(self.current_status_key, self.current_status_text)
        self._apply_row_visuals()

    def _connect_signals(self) -> None:
        self.connect_button.clicked.connect(self._on_connect_clicked)
        self.disconnect_button.clicked.connect(self._on_disconnect_clicked)
        self.start_button.clicked.connect(self._on_start_clicked)
        self.pause_button.clicked.connect(self._on_pause_clicked)
        self.next_stage_button.clicked.connect(self._on_next_stage_clicked)
        self.stop_button.clicked.connect(self._on_stop_clicked)
        self.save_button.clicked.connect(self._on_save_clicked)
        self.load_button.clicked.connect(self._on_load_clicked)
        self.copy_stage_time_button.clicked.connect(self._copy_first_stage_time_to_active_rows)
        self.automation_start_button.clicked.connect(self._on_automation_start_clicked)
        self.automation_add_current_button.clicked.connect(self._on_add_current_plan_to_automation_clicked)
        self.automation_add_file_button.clicked.connect(self._on_add_plan_file_to_automation_clicked)
        self.automation_remove_button.clicked.connect(self._on_remove_automation_rows_clicked)
        self.automation_move_up_button.clicked.connect(lambda: self._move_automation_row(-1))
        self.automation_move_down_button.clicked.connect(lambda: self._move_automation_row(1))
        self.automation_clear_button.clicked.connect(self._on_clear_automation_clicked)
        self.theme_combo.currentIndexChanged.connect(self._on_theme_changed)
        self.keepalive_input.valueChanged.connect(self._update_keepalive_timer_interval)
        self.test_table.itemChanged.connect(self._on_table_item_changed)
        self.automation_table.itemChanged.connect(self._on_automation_table_item_changed)
        self.automation_table.itemSelectionChanged.connect(self._update_automation_buttons)
        self.sequence_controller.log_message.connect(self._append_log)
        self.sequence_controller.state_changed.connect(self._set_status)
        self.sequence_controller.stage_changed.connect(self._on_stage_changed)
        self.sequence_controller.timing_changed.connect(self._on_timing_changed)
        self.sequence_controller.finished.connect(self._on_sequence_finished)
        self.sequence_controller.error_occurred.connect(self._on_sequence_error)
        self.sequence_controller.stage_write_completed.connect(self._on_stage_write_completed)
        self.series_controller.log_message.connect(self._append_log)
        self.series_controller.job_started.connect(self._on_automation_job_started)
        self.series_controller.job_finished.connect(self._on_automation_job_finished)
        self.series_controller.progress_changed.connect(self._on_automation_progress_changed)
        self.series_controller.finished.connect(self._on_automation_finished)
        self.series_controller.stopped.connect(self._on_automation_stopped)
        self.series_controller.error_occurred.connect(self._on_automation_error)
        for checkbox in self.row_checkboxes:
            checkbox.toggled.connect(self._apply_row_visuals)
            checkbox.toggled.connect(self._refresh_summary)
            checkbox.toggled.connect(self._update_time_copy_button)
            checkbox.toggled.connect(self._mark_dirty)
        for checkbox in self.automation_row_checkboxes:
            checkbox.toggled.connect(self._refresh_automation_summary)
            checkbox.toggled.connect(self._update_automation_buttons)
            checkbox.toggled.connect(self._mark_dirty)
        for input_field in self.channel_label_inputs:
            input_field.textChanged.connect(self._on_channel_label_changed)
        for input_field in self.start_value_inputs:
            input_field.textChanged.connect(self._mark_dirty)
        for send_value_button in self.send_value_buttons:
            send_value_button.clicked.connect(self._on_send_start_value_clicked)
        self.pt1_p_input.valueChanged.connect(self._mark_dirty)
        self.pt1_q_input.valueChanged.connect(self._mark_dirty)
        self.pt1_p_register_input.valueChanged.connect(self._mark_dirty)
        self.pt1_q_register_input.valueChanged.connect(self._mark_dirty)

    def _configure_table(self) -> None:
        self._update_table_headers()
        self.test_table.setAlternatingRowColors(True)
        self.test_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.test_table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.test_table.setSortingEnabled(False)
        self.test_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.test_table.verticalHeader().setDefaultSectionSize(30)
        self.test_table.blockSignals(True)
        for row in range(40):
            checkbox = QCheckBox()
            checkbox.setChecked(False)
            self.row_checkboxes.append(checkbox)
            self.test_table.setCellWidget(row, self.COLUMN_ACTIVE, self._wrap_widget(checkbox))

            stage_item = QTableWidgetItem(str(row + 1))
            stage_item.setFlags(stage_item.flags() & ~Qt.ItemIsEditable)
            stage_item.setTextAlignment(Qt.AlignCenter)
            self.test_table.setItem(row, self.COLUMN_STAGE, stage_item)

            for column in range(self.COLUMN_VALUE_START, self.COLUMN_DURATION + 1):
                item = QTableWidgetItem("")
                item.setTextAlignment(Qt.AlignCenter)
                self.test_table.setItem(row, column, item)
        self.test_table.blockSignals(False)

    def _configure_automation_table(self) -> None:
        self.automation_table.setHorizontalHeaderLabels(["Aktiv", "Name", "Quelle", "Wiederholungen", "Notiz"])
        self.automation_table.setAlternatingRowColors(True)
        self.automation_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.automation_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.automation_table.setSortingEnabled(False)
        self.automation_table.verticalHeader().setDefaultSectionSize(30)
        header = self.automation_table.horizontalHeader()
        header.setSectionResizeMode(self.AUTOMATION_COLUMN_ACTIVE, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(self.AUTOMATION_COLUMN_NAME, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(self.AUTOMATION_COLUMN_SOURCE, QHeaderView.Stretch)
        header.setSectionResizeMode(self.AUTOMATION_COLUMN_REPEAT, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(self.AUTOMATION_COLUMN_NOTE, QHeaderView.Stretch)
        self.automation_table.blockSignals(True)
        for row in range(self.AUTOMATION_ROW_COUNT):
            checkbox = QCheckBox()
            checkbox.setChecked(False)
            self.automation_row_checkboxes.append(checkbox)
            self.automation_table.setCellWidget(row, self.AUTOMATION_COLUMN_ACTIVE, self._wrap_widget(checkbox))
            for column in range(self.AUTOMATION_COLUMN_NAME, self.AUTOMATION_COLUMN_NOTE + 1):
                item = QTableWidgetItem("")
                if column == self.AUTOMATION_COLUMN_REPEAT:
                    item.setTextAlignment(Qt.AlignCenter)
                self.automation_table.setItem(row, column, item)
        self.automation_table.blockSignals(False)

    def _wrap_widget(self, widget: QWidget) -> QWidget:
        container = QWidget()
        layout = QHBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setAlignment(Qt.AlignCenter)
        layout.addWidget(widget)
        return container

    def _current_settings(self) -> ConnectionSettings:
        return ConnectionSettings(
            host=self.host_input.text().strip() or "127.0.0.1",
            port=self.port_input.value(),
            slave_id=self.slave_id_input.value(),
            register_format=self.register_format_combo.currentData(),
            keepalive_interval_seconds=self.keepalive_input.value(),
        )

    def _channel_configs(self) -> list[ChannelConfig]:
        return [
            ChannelConfig(
                name=f"S{index}",
                label=channel_label_input.text().strip(),
                start_register=register_input.value(),
                value_type=type_input.currentData(),
            )
            for index, (channel_label_input, register_input, type_input) in enumerate(
                zip(self.channel_label_inputs, self.register_inputs, self.type_inputs),
                start=1,
            )
        ]

    def _keepalive_address(self) -> int:
        return min(spin_box.value() for spin_box in self.register_inputs) if self.register_inputs else 0

    def _start_keepalive_timer(self) -> None:
        self.keepalive_timer.start(max(1, self.keepalive_input.value()) * 1000)
        self._append_log(
            f"Keepalive aktiv: alle {self.keepalive_input.value()} s, Register {self._keepalive_address()}."
        )

    def _stop_keepalive_timer(self) -> None:
        self.keepalive_timer.stop()

    def _update_keepalive_timer_interval(self) -> None:
        if self.keepalive_timer.isActive():
            self.keepalive_timer.start(max(1, self.keepalive_input.value()) * 1000)
            self._append_log(f"Keepalive aktualisiert: Intervall {self.keepalive_input.value()} s.")

    def _on_keepalive_tick(self) -> None:
        if not self.modbus_service.is_connected:
            self.keepalive_timer.stop()
            return
        try:
            response_text = self.modbus_service.keep_alive(address=self._keepalive_address(), count=1)
        except Exception as exc:
            self._append_log(f"Keepalive fehlgeschlagen: {exc}")
            self.series_controller.stop()
            self.sequence_controller.stop()
            self._set_running_widgets(False)
            self.modbus_service.disconnect()
            self._set_connection_widgets(False)
            self._set_status(STATUS_ERROR, "Fehler")
            self._show_error("Verbindung abgebrochen", str(exc))
            return
        self._append_log(f"Keepalive erfolgreich: {response_text}")

    def _build_stage_plan(self) -> list[StageExecution]:
        plan: list[StageExecution] = []
        channel_configs = self._channel_configs()
        for row in range(self.test_table.rowCount()):
            if not self.row_checkboxes[row].isChecked():
                continue
            duration = self._parse_duration(row)
            if duration <= 0:
                continue
            writes: list[ChannelWrite] = []
            skipped_channels: list[str] = []
            for channel_index, channel in enumerate(channel_configs):
                item = self.test_table.item(row, self.COLUMN_VALUE_START + channel_index)
                text = "" if item is None else item.text().strip()
                if not text:
                    skipped_channels.append(channel.name)
                    continue
                try:
                    value = ValueEncoder.coerce_text_value(text, channel.value_type)
                except ValueEncodingError as exc:
                    raise RuntimeError(f"Stufe {row + 1} {channel.name}: {exc}") from exc
                writes.append(
                    ChannelWrite(
                        channel=ChannelConfig(
                            name=channel.name,
                            start_register=channel.start_register,
                            value_type=channel.value_type,
                        ),
                        original_text=text,
                        value=value,
                    )
                )
            if not writes:
                continue
            plan.append(StageExecution(row_index=row, stage_number=row + 1, duration_seconds=duration, writes=writes, skipped_channels=skipped_channels))
        return plan

    def _parse_duration(self, row: int) -> int:
        item = self.test_table.item(row, self.COLUMN_DURATION)
        text = "" if item is None else item.text().strip().replace(",", ".")
        if not text:
            return 0
        try:
            value = float(text)
        except ValueError as exc:
            raise RuntimeError(f"Stufe {row + 1}: Zeit ist keine Zahl") from exc
        if value < 0:
            raise RuntimeError(f"Stufe {row + 1}: Zeit darf nicht negativ sein")
        return int(round(value))

    def _on_connect_clicked(self) -> None:
        settings = self._current_settings()
        while True:
            try:
                self.modbus_service.connect(settings)
            except Exception as exc:
                self.modbus_service.disconnect()
                self._set_connection_widgets(False)
                self._set_status(STATUS_ERROR, "Fehler")
                self._append_log(
                    f"Verbindung fehlgeschlagen: {settings.host}:{settings.port}, Geraete-ID {settings.slave_id}. {exc}"
                )
                if self._show_retry_dialog("Verbindung fehlgeschlagen", str(exc)):
                    continue
                return
            break
        self._set_status(STATUS_CONNECTED, "Verbunden")
        self._append_log(
            "Verbindung hergestellt: "
            f"{settings.host}:{settings.port}, Geraete-ID {settings.slave_id}, "
            f"Format {settings.register_format.label}, Keepalive {settings.keepalive_interval_seconds} s."
        )
        self._set_connection_widgets(True)
        self._start_keepalive_timer()

    def _on_disconnect_clicked(self) -> None:
        self.series_controller.stop()
        self.sequence_controller.stop()
        self._stop_keepalive_timer()
        self.modbus_service.disconnect()
        self._set_connection_widgets(False)
        self._set_status(STATUS_DISCONNECTED, "Getrennt")
        self._append_log("Verbindung getrennt")
        self.last_write_value.setText("Noch keine Schreibaktion")

    def _on_start_clicked(self) -> None:
        if not self.modbus_service.is_connected:
            self._show_error("Keine Verbindung", "Bitte zuerst verbinden.")
            return
        try:
            plan = self._build_stage_plan()
        except RuntimeError as exc:
            self._show_error("Ungültiger Testplan", str(exc))
            return
        if not plan:
            self._append_log("Teststart abgebrochen: keine aktive Stufe mit Zeit und Sollwerten vorhanden.")
            self._show_error("Kein Testplan", "Mindestens eine aktive Stufe mit Zeit > 0 und mindestens einem Sollwert ist erforderlich.")
            return
        self.modbus_service.update_runtime_settings(self._current_settings())
        self._update_keepalive_timer_interval()
        try:
            self.sequence_controller.start(plan)
        except Exception as exc:
            self._set_status(STATUS_ERROR, "Fehler")
            self._show_error("Teststart fehlgeschlagen", str(exc))
            return
        self._set_running_widgets(True)
        self._update_pause_button()

    def _on_automation_start_clicked(self) -> None:
        if self.series_controller.is_running:
            return
        if not self.modbus_service.is_connected:
            self._show_error("Keine Verbindung", "Bitte zuerst verbinden.")
            return
        try:
            jobs = self._build_automation_jobs()
        except RuntimeError as exc:
            self._show_error("Automatisierung ungueltig", str(exc))
            return
        self._capture_automation_snapshot()
        try:
            self.series_controller.start(jobs)
        except Exception as exc:
            self._restore_automation_snapshot()
            self._show_error("Serienlauf fehlgeschlagen", str(exc))
            return
        if self.series_controller.is_running:
            self._set_running_widgets(True)
            self._update_pause_button()

    def _on_pause_clicked(self) -> None:
        if not self.sequence_controller.has_active_plan:
            return
        try:
            if self.sequence_controller.is_paused:
                self.sequence_controller.resume()
                self.last_write_value.setText("Testlauf fortgesetzt")
            else:
                self.sequence_controller.pause()
                self.last_write_value.setText("Testlauf pausiert")
        except Exception as exc:
            self._set_status(STATUS_ERROR, "Fehler")
            self._show_error("Pause/Fortsetzen fehlgeschlagen", str(exc))
            return
        self._set_running_widgets(True)
        self._update_pause_button()

    def _on_stop_clicked(self) -> None:
        self.series_controller.stop()
        self.sequence_controller.stop()
        self._set_running_widgets(False)
        if self.modbus_service.is_connected:
            self._set_status(STATUS_CONNECTED, "Bereit")
        else:
            self._set_status(STATUS_DISCONNECTED, "Getrennt")
        self.last_write_value.setText("Test gestoppt")
        self._append_log("Testlauf manuell gestoppt.")
        self._highlight_current_row(-1)
        self._refresh_summary()

    def _on_save_clicked(self) -> None:
        self._save_as()

    def _on_load_clicked(self) -> None:
        if not self._confirm_discard_unsaved_changes("Vor dem Laden speichern?", "Es gibt ungespeicherte Aenderungen."):
            return
        start_dir = self.current_file_path.parent if self.current_file_path else Path.home()
        file_path, _ = QFileDialog.getOpenFileName(self, "Testplan laden", str(start_dir), "Excel-Dateien (*.xlsx)")
        if not file_path:
            return
        try:
            self._load_from_excel(Path(file_path))
        except Exception as exc:
            self._show_error("Laden fehlgeschlagen", str(exc))
            return
        self.current_file_path = Path(file_path)
        self._set_dirty(False)
        self._update_window_title()
        self._append_log(f"Testplan geladen: {file_path}")

    def _save_as(self) -> bool:
        start_dir = self.current_file_path.parent if self.current_file_path else Path.home()
        default_name = self.current_file_path.name if self.current_file_path else "modbus_testplan.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(self, "Testplan speichern", str(start_dir / default_name), "Excel-Dateien (*.xlsx)")
        if not file_path:
            return False
        path = Path(file_path)
        if path.suffix.lower() != ".xlsx":
            path = path.with_suffix(".xlsx")
        return self._save_plan_to_path(path)

    def _save_plan_to_path(self, path: Path) -> bool:
        try:
            self._save_to_excel(path)
        except Exception as exc:
            self._show_error("Speichern fehlgeschlagen", str(exc))
            return False
        self.current_file_path = path
        self._set_dirty(False)
        self._update_window_title()
        self._append_log(f"Testplan gespeichert: {path}")
        return True

    def _save_to_excel(self, path: Path | BytesIO) -> None:
        workbook = Workbook()
        sheet_connection = workbook.active
        sheet_connection.title = "Verbindung"
        sheet_connection.append(["Feld", "Wert"])
        sheet_connection.append(["host", self.host_input.text().strip() or "127.0.0.1"])
        sheet_connection.append(["port", self.port_input.value()])
        sheet_connection.append(["slave_id", self.slave_id_input.value()])
        sheet_connection.append(["register_format", self.register_format_combo.currentData().value])
        sheet_connection.append(["keepalive_interval_seconds", self.keepalive_input.value()])
        sheet_connection.append(["pt1_p_ms", self.pt1_p_input.value()])
        sheet_connection.append(["pt1_p_start_register", self.pt1_p_register_input.value()])
        sheet_connection.append(["pt1_q_ms", self.pt1_q_input.value()])
        sheet_connection.append(["pt1_q_start_register", self.pt1_q_register_input.value()])

        sheet_channels = workbook.create_sheet("Kanaele")
        sheet_channels.append(["name", "label", "start_register", "value_type", "start_value"])
        for index, (channel_label_input, register_input, type_input, start_value_input) in enumerate(
            zip(self.channel_label_inputs, self.register_inputs, self.type_inputs, self.start_value_inputs),
            start=1,
        ):
            sheet_channels.append(
                [
                    f"S{index}",
                    channel_label_input.text().strip(),
                    register_input.value(),
                    type_input.currentData().value,
                    start_value_input.text().strip(),
                ]
            )

        sheet_plan = workbook.create_sheet("Testplan")
        sheet_plan.append(["active", "stage", "value_1", "value_2", "value_3", "value_4", "duration"])
        for row in range(self.test_table.rowCount()):
            sheet_plan.append(
                [
                    self.row_checkboxes[row].isChecked(),
                    row + 1,
                    *[
                        self._item_text(row, column)
                        for column in range(self.COLUMN_VALUE_START, self.COLUMN_VALUE_START + self.CHANNEL_COUNT)
                    ],
                    self._item_text(row, self.COLUMN_DURATION),
                ]
            )
        sheet_automation = workbook.create_sheet("Automatisierung")
        sheet_automation.append(["active", "name", "source", "repeat_count", "note"])
        for row in range(self.automation_table.rowCount()):
            if self._automation_row_is_empty(row):
                continue
            sheet_automation.append(
                [
                    self.automation_row_checkboxes[row].isChecked(),
                    self._automation_item_text(row, self.AUTOMATION_COLUMN_NAME),
                    self._automation_item_text(row, self.AUTOMATION_COLUMN_SOURCE),
                    self._automation_item_text(row, self.AUTOMATION_COLUMN_REPEAT),
                    self._automation_item_text(row, self.AUTOMATION_COLUMN_NOTE),
                ]
            )
        workbook.save(path)

    def _load_from_excel(self, path: Path | BytesIO, load_automation: bool = True) -> None:
        workbook = load_workbook(path)
        if "Verbindung" not in workbook.sheetnames or "Kanaele" not in workbook.sheetnames or "Testplan" not in workbook.sheetnames:
            raise RuntimeError("Die Excel-Datei muss die Blaetter 'Verbindung', 'Kanaele' und 'Testplan' enthalten.")

        connection_sheet = workbook["Verbindung"]
        connection = {
            str(row[0].value): row[1].value
            for row in connection_sheet.iter_rows(min_row=2, max_col=2)
            if row[0].value is not None
        }
        self.host_input.setText(connection.get("host", "127.0.0.1"))
        self.port_input.setValue(int(connection.get("port", 502)))
        self.slave_id_input.setValue(int(connection.get("slave_id", 1)))
        self.keepalive_input.setValue(int(connection.get("keepalive_interval_seconds", 20)))
        self.pt1_p_input.setValue(int(connection.get("pt1_p_ms", 0)))
        self.pt1_p_register_input.setValue(int(connection.get("pt1_p_start_register", 0)))
        self.pt1_q_input.setValue(int(connection.get("pt1_q_ms", 0)))
        self.pt1_q_register_input.setValue(int(connection.get("pt1_q_start_register", 0)))

        if "register_format" in connection:
            register_format_value = connection.get("register_format", RegisterFormat.BIG.value)
        else:
            register_format_value = RegisterFormat.from_legacy(connection.get("byte_order"), connection.get("word_order")).value
        self._set_combo_value(self.register_format_combo, register_format_value)

        channels_sheet = workbook["Kanaele"]
        channels = []
        for row in channels_sheet.iter_rows(min_row=2, max_col=5, values_only=True):
            if row[0] is None:
                continue
            if len(row) >= 5:
                channels.append(
                    {
                        "name": row[0],
                        "label": row[1],
                        "start_register": row[2],
                        "value_type": row[3],
                        "start_value": row[4],
                    }
                )
            elif len(row) >= 4:
                channels.append(
                    {
                        "name": row[0],
                        "label": row[1],
                        "start_register": row[2],
                        "value_type": row[3],
                        "start_value": "",
                    }
                )
            else:
                channels.append(
                    {
                        "name": row[0],
                        "label": "",
                        "start_register": row[1],
                        "value_type": row[2],
                        "start_value": "",
                    }
                )
        for index, channel in enumerate(channels):
            if index >= len(self.register_inputs):
                break
            self.channel_label_inputs[index].setText("" if channel.get("label") is None else str(channel.get("label", "")))
            self.register_inputs[index].setValue(int(channel.get("start_register", 0)))
            self._set_combo_value(self.type_inputs[index], channel.get("value_type", RegisterValueType.FLOAT32.value))
            self.start_value_inputs[index].setText("" if channel.get("start_value") is None else str(channel.get("start_value", "")))
        self._update_table_headers()

        plan_sheet = workbook["Testplan"]
        rows = []
        for row in plan_sheet.iter_rows(min_row=2, max_col=8, values_only=True):
            if row[1] is None:
                continue
            rows.append(
                {
                    "active": bool(row[0]),
                    "stage": row[1],
                    "values": ["" if value is None else str(value) for value in row[2:6]],
                    "duration": "" if row[6] is None else str(row[6]),
                }
            )
        self.test_table.blockSignals(True)
        for row in range(self.test_table.rowCount()):
            row_data = rows[row] if row < len(rows) else {}
            self.row_checkboxes[row].setChecked(bool(row_data.get("active", False)))
            values = row_data.get("values", ["", "", "", ""])
            for column, value in enumerate(values, start=self.COLUMN_VALUE_START):
                self._set_item_text(row, column, str(value))
            self._set_item_text(row, self.COLUMN_DURATION, str(row_data.get("duration", "")))
        self.test_table.blockSignals(False)
        self._update_time_copy_button()
        self._apply_row_visuals()
        self._refresh_summary()
        if load_automation:
            self._load_automation_from_workbook(workbook)

    def _set_combo_value(self, combo: QComboBox, raw_value: str) -> None:
        normalized_raw = getattr(raw_value, "value", raw_value)
        for index in range(combo.count()):
            data = combo.itemData(index)
            normalized_data = getattr(data, "value", data)
            if data == raw_value or normalized_data == normalized_raw or str(normalized_data) == str(normalized_raw):
                combo.setCurrentIndex(index)
                return

    def _item_text(self, row: int, column: int) -> str:
        item = self.test_table.item(row, column)
        return "" if item is None else item.text().strip()

    def _set_item_text(self, row: int, column: int, value: str) -> None:
        item = self.test_table.item(row, column)
        if item is None:
            item = QTableWidgetItem()
            self.test_table.setItem(row, column, item)
        item.setText(value if value != "None" else "")
        item.setTextAlignment(Qt.AlignCenter)

    def _automation_item_text(self, row: int, column: int) -> str:
        item = self.automation_table.item(row, column)
        return "" if item is None else item.text().strip()

    def _set_automation_item_text(self, row: int, column: int, value: str) -> None:
        item = self.automation_table.item(row, column)
        if item is None:
            item = QTableWidgetItem()
            self.automation_table.setItem(row, column, item)
        item.setText(value if value != "None" else "")
        if column == self.AUTOMATION_COLUMN_REPEAT:
            item.setTextAlignment(Qt.AlignCenter)

    def _automation_row_is_empty(self, row: int) -> bool:
        return (
            not self.automation_row_checkboxes[row].isChecked()
            and not any(
                self._automation_item_text(row, column)
                for column in range(self.AUTOMATION_COLUMN_NAME, self.AUTOMATION_COLUMN_NOTE + 1)
            )
        )

    def _selected_automation_rows(self) -> list[int]:
        return sorted({index.row() for index in self.automation_table.selectionModel().selectedRows()})

    def _next_empty_automation_row(self) -> int:
        for row in range(self.automation_table.rowCount()):
            if self._automation_row_is_empty(row):
                return row
        return -1

    def _has_any_automation_entries(self) -> bool:
        return any(not self._automation_row_is_empty(row) for row in range(self.automation_table.rowCount()))

    def _append_automation_entry(self, name: str, source: str, repeat_count: str = "1", note: str = "") -> None:
        row = self._next_empty_automation_row()
        if row < 0:
            self._show_error("Serienliste voll", "Bitte zuerst bestehende Eintraege entfernen oder die Liste leeren.")
            return
        self._automation_ui_update_in_progress = True
        self.automation_table.blockSignals(True)
        try:
            self.automation_row_checkboxes[row].setChecked(True)
            self._set_automation_item_text(row, self.AUTOMATION_COLUMN_NAME, name)
            self._set_automation_item_text(row, self.AUTOMATION_COLUMN_SOURCE, source)
            self._set_automation_item_text(row, self.AUTOMATION_COLUMN_REPEAT, repeat_count)
            self._set_automation_item_text(row, self.AUTOMATION_COLUMN_NOTE, note)
        finally:
            self.automation_table.blockSignals(False)
            self._automation_ui_update_in_progress = False
        self.automation_table.selectRow(row)
        self._mark_dirty()
        self._refresh_automation_summary()
        self._update_automation_buttons()

    def _clear_automation_row(self, row: int) -> None:
        self.automation_row_checkboxes[row].setChecked(False)
        for column in range(self.AUTOMATION_COLUMN_NAME, self.AUTOMATION_COLUMN_NOTE + 1):
            self._set_automation_item_text(row, column, "")

    def _clear_all_automation_rows(self, mark_dirty: bool = False) -> None:
        self._automation_ui_update_in_progress = True
        self.automation_table.blockSignals(True)
        try:
            for row in range(self.automation_table.rowCount()):
                self._clear_automation_row(row)
        finally:
            self.automation_table.blockSignals(False)
            self._automation_ui_update_in_progress = False
        self.automation_table.clearSelection()
        if mark_dirty:
            self._mark_dirty()

    def _swap_automation_rows(self, source_row: int, target_row: int) -> None:
        source_data = (
            self.automation_row_checkboxes[source_row].isChecked(),
            self._automation_item_text(source_row, self.AUTOMATION_COLUMN_NAME),
            self._automation_item_text(source_row, self.AUTOMATION_COLUMN_SOURCE),
            self._automation_item_text(source_row, self.AUTOMATION_COLUMN_REPEAT),
            self._automation_item_text(source_row, self.AUTOMATION_COLUMN_NOTE),
        )
        target_data = (
            self.automation_row_checkboxes[target_row].isChecked(),
            self._automation_item_text(target_row, self.AUTOMATION_COLUMN_NAME),
            self._automation_item_text(target_row, self.AUTOMATION_COLUMN_SOURCE),
            self._automation_item_text(target_row, self.AUTOMATION_COLUMN_REPEAT),
            self._automation_item_text(target_row, self.AUTOMATION_COLUMN_NOTE),
        )
        self._automation_ui_update_in_progress = True
        self.automation_table.blockSignals(True)
        try:
            self.automation_row_checkboxes[source_row].setChecked(target_data[0])
            self._set_automation_item_text(source_row, self.AUTOMATION_COLUMN_NAME, target_data[1])
            self._set_automation_item_text(source_row, self.AUTOMATION_COLUMN_SOURCE, target_data[2])
            self._set_automation_item_text(source_row, self.AUTOMATION_COLUMN_REPEAT, target_data[3])
            self._set_automation_item_text(source_row, self.AUTOMATION_COLUMN_NOTE, target_data[4])
            self.automation_row_checkboxes[target_row].setChecked(source_data[0])
            self._set_automation_item_text(target_row, self.AUTOMATION_COLUMN_NAME, source_data[1])
            self._set_automation_item_text(target_row, self.AUTOMATION_COLUMN_SOURCE, source_data[2])
            self._set_automation_item_text(target_row, self.AUTOMATION_COLUMN_REPEAT, source_data[3])
            self._set_automation_item_text(target_row, self.AUTOMATION_COLUMN_NOTE, source_data[4])
        finally:
            self.automation_table.blockSignals(False)
            self._automation_ui_update_in_progress = False

    def _move_automation_row(self, direction: int) -> None:
        current_row = self.automation_table.currentRow()
        target_row = current_row + direction
        if current_row < 0 or target_row < 0 or target_row >= self.automation_table.rowCount():
            return
        self._swap_automation_rows(current_row, target_row)
        self.automation_table.selectRow(target_row)
        self._mark_dirty()
        self._refresh_automation_summary()
        self._update_automation_buttons()

    def _refresh_automation_summary(self) -> None:
        active_jobs = 0
        total_repeats = 0
        for row in range(self.automation_table.rowCount()):
            if self._automation_row_is_empty(row):
                continue
            if self.automation_row_checkboxes[row].isChecked():
                active_jobs += 1
                total_repeats += self._automation_repeat_count(row)
        if active_jobs == 0:
            self.automation_summary_label.setText(
                "Noch keine aktiven Serienjobs geplant. Du kannst den aktuellen Plan oder externe Excel-Plaene vormerken."
            )
            return
        self.automation_summary_label.setText(
            f"{active_jobs} Serienjob(s) aktiv, {total_repeats} Wiederholung(en) insgesamt."
        )

    def _automation_repeat_count(self, row: int) -> int:
        text = self._automation_item_text(row, self.AUTOMATION_COLUMN_REPEAT).replace(",", ".")
        if not text:
            return 1
        try:
            return max(1, int(round(float(text))))
        except ValueError:
            return 1

    def _load_automation_from_workbook(self, workbook) -> None:
        self._clear_all_automation_rows()
        if "Automatisierung" not in workbook.sheetnames:
            self._refresh_automation_summary()
            self._update_automation_buttons()
            return
        rows = []
        automation_sheet = workbook["Automatisierung"]
        for row in automation_sheet.iter_rows(min_row=2, max_col=5, values_only=True):
            if not any(value is not None and str(value).strip() for value in row):
                continue
            rows.append(
                {
                    "active": bool(row[0]),
                    "name": "" if row[1] is None else str(row[1]),
                    "source": "" if row[2] is None else str(row[2]),
                    "repeat_count": "1" if row[3] is None else str(row[3]),
                    "note": "" if row[4] is None else str(row[4]),
                }
            )
        self._automation_ui_update_in_progress = True
        self.automation_table.blockSignals(True)
        try:
            for row in range(min(len(rows), self.automation_table.rowCount())):
                row_data = rows[row]
                self.automation_row_checkboxes[row].setChecked(bool(row_data["active"]))
                self._set_automation_item_text(row, self.AUTOMATION_COLUMN_NAME, row_data["name"])
                self._set_automation_item_text(row, self.AUTOMATION_COLUMN_SOURCE, row_data["source"])
                self._set_automation_item_text(row, self.AUTOMATION_COLUMN_REPEAT, row_data["repeat_count"])
                self._set_automation_item_text(row, self.AUTOMATION_COLUMN_NOTE, row_data["note"])
        finally:
            self.automation_table.blockSignals(False)
            self._automation_ui_update_in_progress = False
        self._refresh_automation_summary()
        self._update_automation_buttons()

    def _build_automation_jobs(self) -> list[AutomationJob]:
        jobs: list[AutomationJob] = []
        for row in range(self.automation_table.rowCount()):
            if not self.automation_row_checkboxes[row].isChecked():
                continue
            source_text = self._automation_item_text(row, self.AUTOMATION_COLUMN_SOURCE)
            if not source_text:
                raise RuntimeError(f"Serienjob Zeile {row + 1}: Quelle fehlt")
            source_path = Path(source_text)
            if not source_path.exists():
                raise RuntimeError(f"Serienjob Zeile {row + 1}: Datei nicht gefunden - {source_path}")
            repeat_count = self._automation_repeat_count(row)
            if repeat_count <= 0:
                raise RuntimeError(f"Serienjob Zeile {row + 1}: Wiederholungen muessen groesser als 0 sein")
            name = self._automation_item_text(row, self.AUTOMATION_COLUMN_NAME) or source_path.stem
            note = self._automation_item_text(row, self.AUTOMATION_COLUMN_NOTE)
            jobs.append(
                AutomationJob(
                    row_index=row,
                    name=name,
                    source_path=source_path,
                    repeat_count=repeat_count,
                    note=note,
                )
            )
        if not jobs:
            raise RuntimeError("Es ist kein aktiver Serienjob vorhanden")
        return jobs

    def _capture_automation_snapshot(self) -> None:
        buffer = BytesIO()
        self._save_to_excel(buffer)
        self._automation_snapshot_bytes = buffer.getvalue()
        self._automation_snapshot_file_path = self.current_file_path
        self._automation_snapshot_dirty = self._has_unsaved_changes

    def _restore_automation_snapshot(self) -> None:
        if self._automation_snapshot_bytes is None:
            return
        self._load_from_excel(BytesIO(self._automation_snapshot_bytes), load_automation=True)
        self.current_file_path = self._automation_snapshot_file_path
        if self.modbus_service.is_connected:
            self.modbus_service.update_runtime_settings(self._current_settings())
        self._set_dirty(self._automation_snapshot_dirty)
        self._update_window_title()
        self._automation_snapshot_bytes = None
        self._automation_snapshot_file_path = None
        self._automation_snapshot_dirty = False

    def _start_automation_job(self, job: AutomationJob) -> None:
        self._load_from_excel(job.source_path, load_automation=False)
        if not self.modbus_service.is_connected:
            raise RuntimeError("Keine aktive Modbus-Verbindung")
        active_settings = self.modbus_service.settings
        requested_settings = self._current_settings()
        if active_settings is not None and (
            active_settings.host != requested_settings.host or active_settings.port != requested_settings.port
        ):
            raise RuntimeError(
                "Der Serienlauf unterstuetzt in dieser Version nur Testplaene mit derselben IP-Adresse und demselben Port wie die aktive Verbindung."
            )
        plan = self._build_stage_plan()
        if not plan:
            raise RuntimeError(f"Testplan {job.source_path.name} enthaelt keine aktive Stufe mit Zeit und Sollwerten")
        self.modbus_service.update_runtime_settings(requested_settings)
        self._update_keepalive_timer_interval()
        self.sequence_controller.start(plan)
        self.workspace_tabs.setCurrentIndex(0)

    def _copy_first_stage_time_to_active_rows(self) -> None:
        source_text = self._item_text(0, self.COLUMN_DURATION)
        if not source_text:
            self._update_time_copy_button()
            return
        if not any(self.row_checkboxes[row].isChecked() for row in range(1, self.test_table.rowCount())):
            self._update_time_copy_button()
            return
        previous = self._table_ui_update_in_progress
        self._table_ui_update_in_progress = True
        self.test_table.blockSignals(True)
        try:
            for row in range(1, self.test_table.rowCount()):
                if self.row_checkboxes[row].isChecked():
                    self._set_item_text(row, self.COLUMN_DURATION, source_text)
        finally:
            self.test_table.blockSignals(False)
            self._table_ui_update_in_progress = previous
        self._mark_dirty()
        if not self.sequence_controller.has_active_plan:
            self._refresh_summary()
        self._apply_row_visuals()
        self.last_write_value.setText("Zeit aus Stufe 1 auf aktive Stufen uebernommen")
        self._update_time_copy_button()

    def _on_send_start_value_clicked(self) -> None:
        sender = self.sender()
        try:
            channel_index = self.send_value_buttons.index(sender)
        except ValueError:
            return
        if not self.modbus_service.is_connected:
            self._show_error("Keine Verbindung", "Bitte zuerst verbinden.")
            return
        raw_text = self.start_value_inputs[channel_index].text().strip()
        if not raw_text:
            self._show_error("Startwert fehlt", f"Bitte einen Startwert fuer S{channel_index + 1} eingeben.")
            return
        channel = self._channel_configs()[channel_index]
        try:
            value = ValueEncoder.coerce_text_value(raw_text, channel.value_type)
            result = self.modbus_service.write_channel(channel, value)
        except Exception as exc:
            self._show_error("Startwert senden fehlgeschlagen", str(exc))
            return
        self.last_write_value.setText(f"{channel.name}: Startwert gesendet")
        self._append_log(
            f"Startwert fuer {channel.name} gesendet: Wert {result.original_value} auf Register {result.start_register}."
        )

    def _on_next_stage_clicked(self) -> None:
        if not self.sequence_controller.has_active_plan:
            return
        self.sequence_controller.skip_current_stage()
        self.last_write_value.setText("Zur naechsten Stufe gesprungen")
        self._update_pause_button()

    def _on_table_item_changed(self, item: QTableWidgetItem) -> None:
        if self._table_ui_update_in_progress:
            return
        self._mark_dirty()
        self._table_ui_update_in_progress = True
        try:
            if item.column() >= 2:
                item.setTextAlignment(Qt.AlignCenter)
            if item.row() == 0 and item.column() == self.COLUMN_DURATION:
                self._update_time_copy_button()
            if not self.sequence_controller.has_active_plan:
                self._refresh_summary()
            self._apply_row_visuals()
        finally:
            self._table_ui_update_in_progress = False

    def _on_automation_table_item_changed(self, item: QTableWidgetItem) -> None:
        if self._automation_ui_update_in_progress:
            return
        self._mark_dirty()
        if item.column() == self.AUTOMATION_COLUMN_REPEAT:
            self._automation_ui_update_in_progress = True
            try:
                item.setTextAlignment(Qt.AlignCenter)
            finally:
                self._automation_ui_update_in_progress = False
        self._refresh_automation_summary()
        self._update_automation_buttons()

    def _refresh_summary(self) -> None:
        if self.sequence_controller.has_active_plan:
            return
        try:
            plan = self._build_stage_plan()
        except RuntimeError:
            plan = []
        total = sum(stage.duration_seconds for stage in plan)
        self.total_time_value.setText(self._format_seconds(total))
        self.total_remaining_value.setText(self._format_seconds(total))
        self.stage_remaining_value.setText("00:00:00")
        self.current_stage_value.setText("-")

    def _apply_row_visuals(self) -> None:
        previous = self._table_ui_update_in_progress
        self._table_ui_update_in_progress = True
        try:
            for row in range(self.test_table.rowCount()):
                active = self.row_checkboxes[row].isChecked()
                for column in range(1, self.test_table.columnCount()):
                    item = self.test_table.item(row, column)
                    if item is None:
                        continue
                    flags = item.flags() | Qt.ItemIsSelectable | Qt.ItemIsEnabled
                    if column >= 2:
                        if active and not self.sequence_controller.has_active_plan:
                            item.setFlags(flags | Qt.ItemIsEditable)
                        else:
                            item.setFlags(flags & ~Qt.ItemIsEditable)
                    color = self.active_stage_color if row == self.current_highlight_row else (self.default_row_color if active else self.inactive_row_color)
                    item.setBackground(color)
                    item.setForeground(self.active_row_text_color if active else self.inactive_row_text_color)
        finally:
            self._table_ui_update_in_progress = previous

    def _set_connection_widgets(self, connected: bool) -> None:
        self.connect_button.setEnabled(not connected)
        self.disconnect_button.setEnabled(connected)
        self.host_input.setEnabled(not connected)
        self.port_input.setEnabled(not connected)
        self.slave_id_input.setEnabled(not connected)
        self.register_format_combo.setEnabled(not connected)
        self.keepalive_input.setEnabled(not connected)
        self._update_manual_write_buttons()
        self._update_automation_buttons()

    def _set_running_widgets(self, running: bool) -> None:
        self.start_button.setEnabled(not running)
        self.pause_button.setEnabled(running)
        self.next_stage_button.setEnabled(running)
        self.stop_button.setEnabled(running)
        self.save_button.setEnabled(not running)
        self.load_button.setEnabled(not running)
        self.copy_stage_time_button.setEnabled(False)
        self.test_table.setEnabled(True)
        self.automation_table.setEnabled(not running)
        for spin_box in self.register_inputs:
            spin_box.setEnabled(not running)
        for input_field in self.channel_label_inputs:
            input_field.setEnabled(not running)
        for combo in self.type_inputs:
            combo.setEnabled(not running)
        for pt1_input in (self.pt1_p_input, self.pt1_q_input):
            pt1_input.setEnabled(not running)
        self.disconnect_button.setEnabled(not running and self.modbus_service.is_connected)
        self._update_manual_write_buttons()
        self._update_time_copy_button()
        self._update_pause_button()
        self._update_automation_buttons()
        self._apply_row_visuals()

    def _update_pause_button(self) -> None:
        if self.sequence_controller.is_paused:
            self.pause_button.setText("Fortsetzen")
            self.pause_button.setEnabled(True)
            return
        self.pause_button.setText("Pause")
        self.pause_button.setEnabled(self.sequence_controller.has_active_plan)

    def _update_manual_write_buttons(self) -> None:
        manual_write_enabled = self.modbus_service.is_connected and not self.sequence_controller.has_active_plan
        for input_field in self.start_value_inputs:
            input_field.setEnabled(manual_write_enabled)
        for button in self.send_value_buttons:
            button.setEnabled(manual_write_enabled)

    def _update_time_copy_button(self) -> None:
        has_source_time = bool(self._item_text(0, self.COLUMN_DURATION))
        has_target_rows = any(self.row_checkboxes[row].isChecked() for row in range(1, self.test_table.rowCount()))
        self.copy_stage_time_button.setEnabled(
            has_source_time and has_target_rows and not self.sequence_controller.has_active_plan
        )

    def _update_automation_buttons(self) -> None:
        not_running = not self.sequence_controller.has_active_plan
        selected_rows = self._selected_automation_rows()
        current_row = self.automation_table.currentRow()
        has_entries = self._has_any_automation_entries()
        self.automation_start_button.setEnabled(
            not_running and not self.series_controller.is_running and has_entries and self.modbus_service.is_connected
        )
        self.automation_add_current_button.setEnabled(not_running and not self.series_controller.is_running)
        self.automation_add_file_button.setEnabled(not_running and not self.series_controller.is_running)
        self.automation_remove_button.setEnabled(not_running and not self.series_controller.is_running and bool(selected_rows))
        self.automation_clear_button.setEnabled(not_running and not self.series_controller.is_running and has_entries)
        self.automation_move_up_button.setEnabled(not_running and not self.series_controller.is_running and current_row > 0)
        self.automation_move_down_button.setEnabled(
            not_running
            and not self.series_controller.is_running
            and 0 <= current_row < self.automation_table.rowCount() - 1
        )

    def _set_status(self, status: str, text: str) -> None:
        self.current_status_key = status
        self.current_status_text = text
        self.status_badge.setText(text)
        theme = self.THEMES.get(self.current_theme, self.THEMES["light"])
        status_styles = theme["status_styles"]
        self.status_badge.setStyleSheet(status_styles.get(status, status_styles[STATUS_DISCONNECTED]))

    def _append_log(self, message: str) -> None:
        self.log_panel.appendPlainText(message)
        scrollbar = self.log_panel.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())

    def _on_stage_changed(self, stage_index: int, row_index: int) -> None:
        self.current_stage_value.setText(str(row_index + 1) if row_index >= 0 else "-")
        self._highlight_current_row(row_index)

    def _on_timing_changed(self, total_planned: int, total_remaining: int, stage_remaining: int) -> None:
        self.total_time_value.setText(self._format_seconds(total_planned))
        self.total_remaining_value.setText(self._format_seconds(total_remaining))
        self.stage_remaining_value.setText(self._format_seconds(stage_remaining))

    def _on_sequence_finished(self) -> None:
        if self.series_controller.is_running:
            self.series_controller.plan_finished()
            return
        self._set_running_widgets(False)
        self.last_write_value.setText("Testplan abgeschlossen")
        self._highlight_current_row(-1)

    def _on_sequence_error(self, message: str) -> None:
        if self.series_controller.is_running:
            self.series_controller.plan_failed(message)
            return
        self._set_running_widgets(False)
        self.last_write_value.setText(message)
        self._show_error("Testlauf fehlgeschlagen", message)
        self._highlight_current_row(-1)

    def _on_automation_job_started(
        self,
        row_index: int,
        name: str,
        repeat_index: int,
        repeat_total: int,
        run_index: int,
        run_total: int,
    ) -> None:
        self.automation_table.selectRow(row_index)
        self.workspace_tabs.setCurrentIndex(0)
        self.automation_summary_label.setText(
            f"Serienlauf aktiv: {name} | Wiederholung {repeat_index}/{repeat_total} | Durchgang {run_index}/{run_total}"
        )
        self.last_write_value.setText(f"Serienlauf: {name} ({repeat_index}/{repeat_total})")

    def _on_automation_job_finished(
        self,
        row_index: int,
        name: str,
        repeat_index: int,
        repeat_total: int,
        run_index: int,
        run_total: int,
    ) -> None:
        self.automation_table.selectRow(row_index)
        self.automation_summary_label.setText(
            f"Letzter abgeschlossener Serienjob: {name} | Wiederholung {repeat_index}/{repeat_total} | Durchgang {run_index}/{run_total}"
        )

    def _on_automation_progress_changed(self, completed_runs: int, total_runs: int) -> None:
        if self.series_controller.is_running:
            return
        self.automation_summary_label.setText(
            f"Serienlauf abgeschlossen: {completed_runs}/{total_runs} Durchgaenge beendet."
        )

    def _on_automation_finished(self) -> None:
        self._restore_automation_snapshot()
        self._set_running_widgets(False)
        self._set_status(STATUS_FINISHED, "Serienlauf fertig")
        self.last_write_value.setText("Serienlauf abgeschlossen")
        self._highlight_current_row(-1)
        self._refresh_automation_summary()

    def _on_automation_stopped(self) -> None:
        self._restore_automation_snapshot()
        self._highlight_current_row(-1)
        self._refresh_automation_summary()

    def _on_automation_error(self, message: str) -> None:
        self._restore_automation_snapshot()
        self._set_running_widgets(False)
        self._set_status(STATUS_ERROR, "Fehler")
        self.last_write_value.setText(message)
        self._highlight_current_row(-1)
        self._refresh_automation_summary()
        self._show_error("Serienlauf fehlgeschlagen", message)

    def _on_stage_write_completed(self, stage_number: int, message: str) -> None:
        self.last_write_value.setText(f"Stufe {stage_number}: {message}")

    def _highlight_current_row(self, row_index: int) -> None:
        self.current_highlight_row = row_index
        self._apply_row_visuals()

    def _on_theme_changed(self) -> None:
        self._apply_theme(self.theme_combo.currentData())

    def _on_channel_label_changed(self) -> None:
        self._update_table_headers()
        self._mark_dirty()

    def _on_add_current_plan_to_automation_clicked(self) -> None:
        if self.current_file_path is None:
            self._show_error("Testplan zuerst speichern", "Bitte den aktuellen Testplan zuerst als Excel-Datei speichern.")
            return
        source = str(self.current_file_path)
        name = self.current_file_path.stem
        note = "Aus aktueller Konfiguration uebernommen"
        self._append_automation_entry(name=name, source=source, repeat_count="1", note=note)

    def _on_add_plan_file_to_automation_clicked(self) -> None:
        start_dir = self.current_file_path.parent if self.current_file_path else Path.home()
        file_path, _ = QFileDialog.getOpenFileName(self, "Plan fuer Serienlauf auswaehlen", str(start_dir), "Excel-Dateien (*.xlsx)")
        if not file_path:
            return
        selected_path = Path(file_path)
        self._append_automation_entry(
            name=selected_path.stem,
            source=str(selected_path),
            repeat_count="1",
            note="Externer Plan fuer Serienlauf",
        )

    def _on_remove_automation_rows_clicked(self) -> None:
        selected_rows = self._selected_automation_rows()
        if not selected_rows:
            return
        self._automation_ui_update_in_progress = True
        self.automation_table.blockSignals(True)
        try:
            for row in selected_rows:
                self._clear_automation_row(row)
        finally:
            self.automation_table.blockSignals(False)
            self._automation_ui_update_in_progress = False
        self._mark_dirty()
        self._refresh_automation_summary()
        self._update_automation_buttons()

    def _on_clear_automation_clicked(self) -> None:
        if not self._has_any_automation_entries():
            return
        self._clear_all_automation_rows(mark_dirty=True)
        self._refresh_automation_summary()
        self._update_automation_buttons()

    def _update_table_headers(self) -> None:
        headers = ["Aktiv", "Stufe"]
        for index, input_field in enumerate(self.channel_label_inputs, start=1):
            label = input_field.text().strip()
            headers.append(f"Sollwert {index}" if not label else f"Sollwert {index}\n{label}")
        headers.append("Zeit [s]")
        self.test_table.setHorizontalHeaderLabels(headers)

    def _format_seconds(self, total_seconds: int) -> str:
        seconds = max(0, int(total_seconds))
        return f"{seconds // 3600:02d}:{(seconds % 3600) // 60:02d}:{seconds % 60:02d}"

    def _show_error(self, title: str, message: str) -> None:
        QMessageBox.critical(self, title, message)

    def _show_retry_dialog(self, title: str, message: str) -> bool:
        answer = QMessageBox.question(
            self,
            title,
            f"{message}\n\nMoechtest du es erneut versuchen?",
            QMessageBox.Retry | QMessageBox.Cancel,
            QMessageBox.Retry,
        )
        return answer == QMessageBox.Retry

    def closeEvent(self, event: QCloseEvent) -> None:  # type: ignore[override]
        if not self._confirm_discard_unsaved_changes("Vor dem Schliessen speichern?", "Es gibt ungespeicherte Aenderungen."):
            event.ignore()
            return
        self.keepalive_timer.stop()
        self.series_controller.stop()
        self.sequence_controller.stop()
        self.modbus_service.disconnect()
        super().closeEvent(event)

    def _mark_dirty(self) -> None:
        self._set_dirty(True)

    def _set_dirty(self, dirty: bool) -> None:
        if self._has_unsaved_changes == dirty:
            return
        self._has_unsaved_changes = dirty
        self._update_window_title()

    def _update_window_title(self) -> None:
        file_name = self.current_file_path.name if self.current_file_path else "Unbenannt"
        dirty_marker = " *" if self._has_unsaved_changes else ""
        self.setWindowTitle(f"{self.DEFAULT_WINDOW_TITLE} - {file_name}{dirty_marker}")

    def _confirm_discard_unsaved_changes(self, title: str, message: str) -> bool:
        if not self._has_unsaved_changes:
            return True
        answer = QMessageBox.question(
            self,
            title,
            f"{message}\n\nMoechtest du zuerst speichern?",
            QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel,
            QMessageBox.Save,
        )
        if answer == QMessageBox.Cancel:
            return False
        if answer == QMessageBox.Save:
            return self._save_as()
        return True


def run() -> int:
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    window = ModbusMainWindow()
    window.show()
    return app.exec_()

